import os
import requests
from bs4 import BeautifulSoup
import openpyxl
import time
import json
from fpdf import FPDF
from Bio import Entrez
import re
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
import dateutil.parser as dparser

Entrez.email = "eric7lee87@gmail.com" 

def pubmed_search_ids(query):
    '''
    Parameters:
    ----------------
    query: the query of the pubmed search

    Returns:
    ----------------
    ids_list: list of ids that query search returns
    '''

    id_list = []
    # the url for PubMed query search
    search_ids_url = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi?term="

    # the response from the url
    ids_response = requests.get(search_ids_url + query)

    # sleep for 0.4 seconds to not get blocked
    time.sleep(0.4)

    # the beautiful soup of the url
    ids_soup = BeautifulSoup(ids_response.text, features="xml")

    # the list of ids in xml format
    ids_xml_list = ids_soup.find_all('Id')

    for id_tag in ids_xml_list:
        if id_tag is not None:  
            # Extract the text content of the Id tag
            id_num = int(id_tag.text)
            id_list.append(id_num)

    return id_list  # Return id_list instead of ids_list

path = "../data/papers.xlsx"
def extract_metadata(path):
    '''
    Parameters:
    ----------------
    path: the path of the excel search

    Returns:
    ----------------
    articles: dictionary of articles where the key is the sheet name and the value is a list of articles
    '''
    articles = {}

    # Define variable to load the dataframe
    dataframe = openpyxl.load_workbook(path)

    # iterates through each sheet and adds each title into list
    for i in range(len(dataframe.sheetnames)):
        sheetname = dataframe.sheetnames[i]
        curr_sheet = dataframe._sheets[i]
        for i in range(1, curr_sheet.max_row):
            title = curr_sheet.cell(row = i, column = 8).value #H
            authors = curr_sheet.cell(row = i, column = 5).value #E
            pmid = curr_sheet.cell(row = i, column = 3).value #C
            date = curr_sheet.cell(row = i, column = 13).value #M
            abstract = curr_sheet.cell(row = i, column = 11).value #K
            mesh = curr_sheet.cell(row = i, column = 14).value #N
            keywords = curr_sheet.cell(row = i, column = 17).value #Q
            if mesh is not None:
                mesh = mesh.split()
            if keywords is not None:
                keywords = keywords.split()
            if title is not None:
                articles[title] = [title, authors, pmid, date, abstract, mesh, keywords]

    return articles

def pubmed_search_articles(search):
    '''
    Parameters:
    ----------------
    search: the search query on pubmed

    Returns:
    ----------------
    id_list: the list of the urls
    '''
    id_list = []

    # the url of the article
    url = "https://pubmed.ncbi.nlm.nih.gov/?term=" + search + "&filter=simsearch3.fft"

    # the response from the url
    response = requests.get(url)

    # sleep for 0.4 seconds to not get blocked
    time.sleep(0.4)

    # the beautiful soup of the url
    soup = BeautifulSoup(response.text, features="lxml")

    try:
        # Find all 'a' tags with class 'docsum-title'
        a_tags = soup.find_all('a', class_='docsum-title')

        # Extract and print the data-article-id for each tag
        for i, tag in enumerate(a_tags, 1):
            data_article_id = tag.get('data-article-id')
            id_list.append(data_article_id)

        return id_list
    except:
        return None



def pubmed_get_abstracts(id):
    '''
    Parameters:
    ----------------
    id: the id of the article

    Returns:
    ----------------
    abstract: the abstract of the article
    '''
    abstract = ""

    # the url of the article
    url = "https://pubmed.ncbi.nlm.nih.gov/" + str(id)

    # the response from the url
    response = requests.get(url)

    # sleep for 0.4 seconds to not get blocked
    time.sleep(0.4)

    # the beautiful soup of the url
    soup = BeautifulSoup(response.text, features="lxml")
    
    try:
        # finds the abstract content
        abstract_soup = soup.find("div", {"class": "abstract-content selected"})
        paragraph_soup = abstract_soup.find_all("p")

        for p in paragraph_soup:
            abstract = abstract + " " + p.get_text(strip=True)

        return abstract
    except:
        return None

def pubmed_extract_metadata(id):
    '''
    Parameters:
    ----------------
    id: the id of the article

    Returns:
    ----------------
    metadata: the metadata of the article
    '''
    metadata = []

    url = "https://pubmed.ncbi.nlm.nih.gov/" + str(id)

    # the response from the url
    response = requests.get(url)

    # sleep for 0.4 seconds to not get blocked
    time.sleep(0.4)

    # the beautiful soup of the url
    soup = BeautifulSoup(response.text, features="lxml")

    title = ""
    authors = []
    date = ""
    # Find the h1 tag with class "heading-title"
    h1_tag = soup.find('h1', class_='heading-title')

    if h1_tag:
        # Extract the text, strip whitespace
        title = h1_tag.get_text(strip=True)

    # Find all 'a' tags with class "full-name"
    author_tags = soup.find_all('a', class_='full-name')

    # Extract and print the author names
    for tag in author_tags:
        author_name = tag.get_text(strip=True)
        authors.append(author_name)

    # Find the span tag with class "cit"
    cit_span = soup.find('span', class_='cit')

    if cit_span:
        # Extract the text content
        date = cit_span.get_text(strip=True)
        match = re.match(r'(\d{4} \w{3} \d{2})', date)
        if match:
            date = match.group(1)

    metadata = [title, authors, date, id]
    return metadata



def get_full_text_from_search_results(search_results_xml):
    # batch_size = 100
    # for start in range(0, int(search_results["Count"]), batch_size):
    #     end = min(int(search_results["Count"]), start + batch_size)
    #     handle = Entrez.efetch(db="pmc", rettype="full", retmode="xml", 
    #                         retstart=start, retmax=batch_size,
    #                         webenv=search_results["WebEnv"], 
    #                         query_key=search_results["QueryKey"])
    #     data = handle.read()
        soup = BeautifulSoup(search_results_xml, features="xml")
        # body_soup = soup.find("body")
        p_soup = soup.find_all("p")
        full_text = ""
        for p in p_soup:
            full_text = full_text + p.text
        return full_text


def create_pdf(article_dict, output_folder):
    filename = f"{article_dict['pmid']}.pdf"
    filepath = os.path.join(output_folder, filename)
    
    pdf = FPDF()
    pdf.add_page()
    
    # Set font for title
    pdf.set_font("Arial", 'B', size=16)
    
    # Add title
    pdf.multi_cell(0, 10, txt=article_dict['title'], align='C')
    pdf.ln(10)
    
    # Set font for normal text
    pdf.set_font("Arial", size=12)
    
    # Add authors
    pdf.multi_cell(0, 10, txt=f"Authors: {article_dict['authors']}", align='L')
    pdf.ln(5)
    
    # Add publication date
    pdf.multi_cell(0, 10, txt=f"Publication Date: {article_dict['publication_date']}", align='L')
    pdf.ln(10)
    
    # Add abstract
    if article_dict['abstract'] is not None:
        pdf.set_font("Arial", 'B', size=14)
        pdf.cell(0, 10, txt="Abstract:", ln=1)
        pdf.set_font("Arial", size=12)
        pdf.multi_cell(0, 10, txt=article_dict['abstract'])
        pdf.ln(10)
    else:
        pdf.set_font("Arial", 'B', size=14)
        pdf.cell(0, 10, txt="Abstract:", ln=1)
        pdf.set_font("Arial", size=12)
        pdf.multi_cell(0, 10, txt='Not Available')
        pdf.ln(10)

    pdf.output(filepath)
    print(f"Created PDF for article {article_dict['pmid']}")


def create_large_pdf(dict_list, output_folder, filename, include_fulltexts):
    filepath = os.path.join(output_folder, f"{filename}.pdf")
    doc = SimpleDocTemplate(filepath, pagesize=letter)
    styles = getSampleStyleSheet()
    story = []

    # Create a new style for justified text
    justified_style = ParagraphStyle(name='Justified', parent=styles['Normal'], alignment=TA_JUSTIFY)

    # Add filename as title
    story.append(Paragraph(filename.capitalize(), styles['Title']))
    story.append(Spacer(1, 12))

    for article in dict_list:
        # Add title
        story.append(Paragraph(article['title'], styles['Heading1']))
        story.append(Spacer(1, 12))

        # Add metadata
        metadata = [
            f"Authors: {article['authors']}",
            f"PMID: {article['pmid']}"
        ]
        for item in metadata:
            story.append(Paragraph(item, styles['Normal']))
            story.append(Spacer(1, 6))

        story.append(Spacer(1, 12))

        # Add abstract
        story.append(Paragraph("Abstract", styles['Heading2']))
        abstract_text = article['abstract'] if article['abstract'] is not None else 'Not Available'
        abstract_text = clean_latex(abstract_text)
        story.append(Paragraph(abstract_text, justified_style))
        story.append(Spacer(1, 12))

        # Add full text if required
        if include_fulltexts and article['full_text'] is not None and not article['full_text'] == "":
            story.append(Paragraph("Full Text", styles['Heading2']))
            full_text = article['full_text'] #if article['full_text'] is not None else 'Not Available'
            full_text = clean_latex(full_text)
            
            # Split the full text into sections
            sections = re.split(r'(\n\s*[A-Z][A-Za-z\s]+:)', full_text)
            for i in range(0, len(sections), 2):
                if i+1 < len(sections): 
                    # Add section header
                    story.append(Paragraph(sections[i].strip(), styles['Heading3']))
                    # Add section content
                    story.append(Paragraph(sections[i+1].strip(), justified_style))
                else:
                    # If there's an odd number of sections, add the last one
                    story.append(Paragraph(sections[i].strip(), justified_style))
                story.append(Spacer(1, 12))

        # # Add page break after each article
        # story.append(PageBreak())

    doc.build(story)

def clean_latex(text):
    # Remove LaTeX commands
    text = re.sub(r'\\[a-zA-Z]+(\{[^}]*\})?', '', text)
    # Replace LaTeX special characters
    replacements = {
        '\\&': '&',
        '\\%': '%',
        '\\$': '$',
        '\\#': '#',
        '\\_': '_',
        '\\{': '{',
        '\\}': '}',
        '~': ' ',
        '``': '"',
        "''": '"'
    }
    for latex, char in replacements.items():
        text = text.replace(latex, char)
    # Remove any remaining backslashes
    text = text.replace('\\', '')
    return text

############################################## FULL TEXT ACCESSOR TESTING #################################################
# from Bio import Entrez

# Entrez.email = "eric7lee87@gmail.com" 

# search_query = 'Fibrin Sealants in Dura Sealing: A Systematic Literature Review' + '[Title]'# AND ' + str(author_query) + '[Author]' # + ' AND medline[sb] AND "open access"[filter]'
# search_results = Entrez.read(Entrez.esearch(db="pmc", term=search_query, retmax=10, usehistory="y"))
# batch_size = 100
# for start in range(0, int(search_results["Count"]), batch_size):
#     end = min(int(search_results["Count"]), start + batch_size)
#     handle = Entrez.efetch(db="pmc", rettype="full", retmode="xml", 
#                         retstart=start, retmax=batch_size,
#                         webenv=search_results["WebEnv"], 
#                         query_key=search_results["QueryKey"])
#     data = handle.read()
#     soup = BeautifulSoup(data, features="xml")
#     # body_soup = soup.find("body")
#     p_soup = soup.find_all("p")
#     full_text = ""
#     for p in p_soup:
#         full_text = full_text + p.text
#     print(full_text)
#     # handle.close()

############ CODE THAT CREATES KEYWORD DICTIONARY ###########################

# articles_accessed = 0
# articles = extract_metadata(path)
# keyword_dict = {}
# for article in articles:
#     articles_accessed += 1
#     title = article[0]
#     pmid = article[2]
#     mesh = article[5]
#     keywords = article[6]

#     if mesh is not None and keywords is not None:
#         combined = mesh + keywords
#         combined = list(set(combined))
#     elif mesh is None and keywords is None:
#         continue
#     elif mesh is None and keywords is not None:
#         combined = keywords
#     elif keywords is None and mesh is not None:
#         combined = mesh

#     if combined is not None:
#         for word in combined:
#             word = word.lower()
#             for letter in word:
#                 if letter in '*)]";,':
#                     word = word.replace(letter,'')
#                 elif letter in "[(/":
#                     if word.index(letter) == 0:
#                         word = word.replace(letter,'')
#                     else:
#                         word = word[0:word.index(letter)]
#             if word in keyword_dict.keys():
#                 if title not in keyword_dict[word]:
#                     keyword_dict[word].append(title)
#             else:
#                 keyword_dict[word] = [title]

# json_file_path = "../data/title_dictionary.json"

# if not os.path.exists(json_file_path):
#     with open(json_file_path, 'w') as f:
#         json.dump([], f)

# with open(json_file_path, 'w') as f:
#     json.dump(keyword_dict, f, indent=2)

############### CODE THAT DOWNLOADS PDFS FROM EXCEL SHEET ################
def download_pdf_excel(word, include_fulltexts):
    # getting the titles from the excel sheet
    path = "../data/papers.xlsx"
    articles = extract_metadata(path)

    output_folder = "..\\pdfs\\" + word
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)      
        
    dict_list = []

    articles_found = 0

    ## getting the relevant PMIDs from json

    file = open('../data/title_dictionary.json')
    data = json.load(file)
    titles = data[word.lower()]
    
    for title in titles:
        if title in articles.keys():
            metadata = articles[title]
            authors = metadata[1]
            pmid = metadata[2]
            date = metadata[3]
            abstract = metadata[4]
            authors_list = [line.strip() for line in authors.strip().split('\n\n')]
            author_query = authors_list[0]
            mesh = metadata[5]
            keywords = metadata[6]
            fulltext = ""

        if title == 'TITLE':
            continue

        article_dict = get_article_dict(metadata)
        
        dict_list.append(article_dict)
        articles_found = articles_found + 1
        print("Article " + str(pmid) + " found.")

    print("Creating large pdf...")
    create_large_pdf(dict_list, output_folder, word, include_fulltexts)
    print("PDF created.")

############### CODE THAT DOWNLOADS PDFS NOT IN EXCEL SHEET ################
def download_pdf_not_excel(word, theme, include_fulltexts):
    # getting the titles from the excel sheet
    path = "../data/papers.xlsx"
    articles = extract_metadata(path)

    output_folder = "..\\pdfs\\" + word
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)      
        
    dict_list = []

    articles_found = 0

    ## getting the relevant article titles from json
    file = open('../data/dictionary.json')
    data = json.load(file)

    ## checks if article is in excel sheet
    if word in data.keys():
        pmids = data[word.lower()]
        for article in articles:
            try:
                title = article[0]
                authors = article[1]
                pmid = article[2]
                date = article[3]
                abstract = article[4]
                authors_list = [line.strip() for line in authors.strip().split('\n\n')]
                author_query = authors_list[0]
                mesh = article[5]
                keywords = article[6]
                fulltext = ""
                metadata = [title, authors, pmid, date, abstract]

                if pmid in pmids:

                    if title == 'TITLE':
                        continue

                    article_dict = get_article_dict(metadata)

                    dict_list.append(article_dict)
                    articles_found = articles_found + 1
                    print(str(articles_found) + "/" + str(len(pmids)) + " articles found")

            except Exception as e:
                print("Error: " + str(e))

    ## searches PMC for related articles
    id_list = pubmed_search_articles(word + " " + theme)
    for id in id_list:
        if len(dict_list) >= 10:
            break
        abstract = pubmed_get_abstracts(id)
        metadata = pubmed_extract_metadata(id)
        title = metadata[0]
        authors = metadata[1]
        date = metadata[2]
        metadata.append(abstract)

        article_dict = get_article_dict(metadata)

        dict_list.append(article_dict)
        articles_found = articles_found + 1
        # print(str(articles_found) + "/" + str(len(id_list)) + " articles found")
        print("Article " + str(id) + " found.")

    print("Creating large pdf...")
    create_large_pdf(dict_list, output_folder, word, include_fulltexts)
    print("PDF created.")

######################## CODE USED FOR TESTING EXTRACTION OF FULLTEXTS #####################################

def new_get_full_text(search_results):
    batch_size = 100
    for start in range(0, int(search_results["Count"]), batch_size):
        end = min(int(search_results["Count"]), start + batch_size)
        handle = Entrez.efetch(db="pmc", rettype="full", retmode="xml", 
                            retstart=start, retmax=batch_size,
                            webenv=search_results["WebEnv"], 
                            query_key=search_results["QueryKey"])
        data = handle.read()

        def extract_xref_p_values(text):
            # Create a BeautifulSoup object
            soup = BeautifulSoup(text, 'xml')
            
            # Find all xref tags with ref-type="bibr"
            ps = soup.find_all('p')
            
            # Extract the text content (p value) from each xref
            p_values = [p.text for p in ps]
            
            return p_values


        fulltext = extract_xref_p_values(data)
        return fulltext

##########################  MAIN CODE THAT RUNS THE DOWNLOADING PDFS (CASE HANDLING)  ################################

# file = open('../data/dictionary.json')
# data = json.load(file)
# keyword = input("Enter your keyword:\n")
# theme = "Fibren Glue"

# if keyword not in data.keys() or len(data[keyword]) < 3:
#     ## NOT ENOUGH ARTICLES
#     response = input("Not enough articles available in database. Fill the rest with other articles from PMC?\n")
#     if response.lower() == 'yes' or response.lower() == 'y':
#         download_pdf_not_excel(keyword, theme, False)
#     else:
#         print("Sorry, not enough PDFs.")
# else:
#     download_pdf_excel(keyword, False)


##################### FUNCTION THAT GETS ARTICLE METADATA IN DICT FORMAT #########################################

def get_article_dict(metadata):
    title = metadata[0]
    authors = metadata[1]
    date = metadata[2]
    pmid = metadata[3]
    abstract = metadata[4]
    if type(authors) is not list:
        authors_list = [line.strip() for line in authors.strip().split('\n\n')]
    else:
        authors_list = authors
    author_query = authors_list[0]
    fulltext = ""

    search_query = str(title) + '[Title]'# AND ' + str(author_query) + '[Author]' # + ' AND medline[sb] AND "open access"[filter]'
    search_results = Entrez.read(Entrez.esearch(db="pmc", term=search_query, retmax=10, usehistory="y"))
    time.sleep(1)
    success = False

    if int(search_results["Count"]) == 1: ## if searching only for title worked
        success = True

    elif int(search_results["Count"]) > 1: ## if too many queries popped up
        print("Too many searches: " + str(search_results["Count"]))
        search_query = str(title) + '[Title] AND ' + str(author_query) + '[Author]' ## try searching with author
        search_results = Entrez.read(Entrez.esearch(db="pmc", term=search_query, retmax=10, usehistory="y"))
        time.sleep(1)
        if int(search_results["Count"]) > 1: ## searching with author did not work
            print("Too many searches: " + str(search_results["Count"]))
            for i in range(1, len(authors_list)):
                author_query = " OR "+ author_query
            search_query = str(title) + '[Title] AND ' + str(author_query) + '[Author]' ## try searching with authors
            search_results = Entrez.read(Entrez.esearch(db="pmc", term=search_query, retmax=10, usehistory="y"))
            time.sleep(1)
            if int(search_results["Count"]) > 1: ## searching with authors did not work
                print("Still too many searches: " + str(str(search_results["Count"])))
                search_query = str(title) + '[Title] AND ' + str(author_query) + '[Author] AND medline[sb] AND "open access"[filter]' ## try searching with filters
                search_results = Entrez.read(Entrez.esearch(db="pmc", term=search_query, retmax=10, usehistory="y"))
                time.sleep(1)
                if int(search_results["Count"]) > 1: ## searching with filter did not work
                    print("Still too many searches: " + str(str(search_results["Count"])))
                    search_query = str(title) + '[Title] AND ' + str(author_query) + '[Author] AND medline[sb] AND "open access"[filter]' ## try searching with dates
                    search_results = Entrez.read(Entrez.esearch(db="pmc", term=search_query, retmax=10, usehistory="y", mindate = date, maxdate = date))
                    time.sleep(1)
                    if int(search_results["Count"]) > 1: ## searching with dates did not work
                        success = False
                    elif int(search_results["Count"]) == 1: ## searching with dates worked
                        success = True
                elif int(search_results["Count"]) == 1: ## searching with filters worked
                    success = True 
            elif int(search_results["Count"]) == 1: ## searching with authors worked
                success = True 
        elif int(search_results["Count"]) == 1: ## searching with author worked
            success = True 
    else:
        print("Article not found")

    if success:
        fulltext = new_get_full_text(search_results)
        print("Is Success")
    else:
        print("Full text not available")

    ## Finding abstracts
    if abstract == "" or abstract is None:
        id_list = pubmed_search_ids(title)
        if len(id_list) == 1: # checking to see if only one query came back
            abstract = pubmed_get_abstracts(id_list[0])

    ## Create article dictionary
    article_dict = {
        "pmid": pmid,
        "title": title,
        "authors": ", ".join(authors_list),
        "abstract": abstract,
        "full_text": str(fulltext),
        "publication_date": date
    }
    return article_dict

###########################################################################################################################

output_folder = "..\\pdfs"
word = "test"
# title = "Management of chondral and osteochondral lesions of the hip : A comprehensive review"
# authors = ["Rajesh Itha", "Raju Vaishya"]
# pmid = "37815635"
# date = "2023"
# abstract = 'Chondral and osteochondral lesions encompass several acute or chronic defects of the articular cartilage and/or subchondral bone. These lesions can result from several different diseases and injuries, including osteochondritis dissecans, osteochondral defects, osteochondral fractures, subchondral bone osteonecrosis, and insufficiency fractures. As the cartilage has a low capacity for regeneration and self-repair, these lesions can progress to osteoarthritis. This study provides a comprehensive overview of the subject matter that it covers. PubMed, Scopus and Google Scholar were accessed using the following keywords: "chondral lesions/defects of the femoral head", "chondral/cartilage lesions/defects of the acetabulum", "chondral/cartilage lesions/defects of the hip", "osteochondral lesions of the femoral head", "osteochondral lesions of the acetabulum", "osteochondral lesions of the hip", "osteochondritis dissecans," "early osteoarthritis of the hip," and "early stage avascular necrosis". Hip osteochondral injuries can cause significant damage to the articular surface and diminish the quality of life. It can be difficult to treat such injuries, especially in patients who are young and active. Several methods are used to treat chondral and osteochondral injuries of the hip, such as mesenchymal stem cells and cell-based treatment, surgical repair, and microfractures. Realignment of bony anatomy may also be necessary for optimal outcomes. Despite several treatments being successful, there is a lack of head-to-head comparisons and large sample size studies in the current literature. Additional research will be required to provide appropriate clinical recommendations for treating chondral/osteochondral injuries of the hip joint.'
# authors_list = ["Rajesh Itha", "Raju Vaishya"]
# author_query = authors_list[0]
# fulltext = ""

# title = "Fibrin Sealants in Dura Sealing: A Systematic Literature Review"
# authors = "Felice Esposito"
# pmid = "27119993"
# date = "2016"
# abstract = 'Chondral and osteochondral lesions encompass several acute or chronic defects of the articular cartilage and/or subchondral bone. These lesions can result from several different diseases and injuries, including osteochondritis dissecans, osteochondral defects, osteochondral fractures, subchondral bone osteonecrosis, and insufficiency fractures. As the cartilage has a low capacity for regeneration and self-repair, these lesions can progress to osteoarthritis. This study provides a comprehensive overview of the subject matter that it covers. PubMed, Scopus and Google Scholar were accessed using the following keywords: "chondral lesions/defects of the femoral head", "chondral/cartilage lesions/defects of the acetabulum", "chondral/cartilage lesions/defects of the hip", "osteochondral lesions of the femoral head", "osteochondral lesions of the acetabulum", "osteochondral lesions of the hip", "osteochondritis dissecans," "early osteoarthritis of the hip," and "early stage avascular necrosis". Hip osteochondral injuries can cause significant damage to the articular surface and diminish the quality of life. It can be difficult to treat such injuries, especially in patients who are young and active. Several methods are used to treat chondral and osteochondral injuries of the hip, such as mesenchymal stem cells and cell-based treatment, surgical repair, and microfractures. Realignment of bony anatomy may also be necessary for optimal outcomes. Despite several treatments being successful, there is a lack of head-to-head comparisons and large sample size studies in the current literature. Additional research will be required to provide appropriate clinical recommendations for treating chondral/osteochondral injuries of the hip joint.'
# authors_list = ["Felice Esposito"]
# author_query = authors_list[0]
# fulltext = ""

title = "Endoscopic tissue shielding method with polyglycolic acid sheets and fibrin glue to cover wounds after colorectal endoscopic submucosal dissection (with video)"
authors = "Tsuji Y."
pmid = "603253584"
date = "2014"
abstract = 'Chondral and osteochondral lesions encompass several acute or chronic defects of the articular cartilage and/or subchondral bone. These lesions can result from several different diseases and injuries, including osteochondritis dissecans, osteochondral defects, osteochondral fractures, subchondral bone osteonecrosis, and insufficiency fractures. As the cartilage has a low capacity for regeneration and self-repair, these lesions can progress to osteoarthritis. This study provides a comprehensive overview of the subject matter that it covers. PubMed, Scopus and Google Scholar were accessed using the following keywords: "chondral lesions/defects of the femoral head", "chondral/cartilage lesions/defects of the acetabulum", "chondral/cartilage lesions/defects of the hip", "osteochondral lesions of the femoral head", "osteochondral lesions of the acetabulum", "osteochondral lesions of the hip", "osteochondritis dissecans," "early osteoarthritis of the hip," and "early stage avascular necrosis". Hip osteochondral injuries can cause significant damage to the articular surface and diminish the quality of life. It can be difficult to treat such injuries, especially in patients who are young and active. Several methods are used to treat chondral and osteochondral injuries of the hip, such as mesenchymal stem cells and cell-based treatment, surgical repair, and microfractures. Realignment of bony anatomy may also be necessary for optimal outcomes. Despite several treatments being successful, there is a lack of head-to-head comparisons and large sample size studies in the current literature. Additional research will be required to provide appropriate clinical recommendations for treating chondral/osteochondral injuries of the hip joint.'
authors_list = ["Rajesh Itha", "Raju Vaishya"]
author_query = authors_list[0]
fulltext = ""

# metadata = [title, authors, pmid, date, abstract]
# article_dict = get_article_dict(metadata)
# print(article_dict)

# print("Creating large pdf...")
# create_large_pdf([article_dict], output_folder, word, True)
# print("PDF created.")


download_pdf_not_excel("doctor", "fibren glue", True)